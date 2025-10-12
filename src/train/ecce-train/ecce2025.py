#!/usr/bin/env python
"""ECCE-2025-1779.ipynb
Author : Emmanuel Havugimana
Related: ECCE-2025 paper 17xx

Generator script

"""

"""# Run the model"""
import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import scipy
import torch
from torch import Tensor
import torch.nn as nn
import torch.nn.functional as F
import torch.optim as optim
import random
import datetime
import shutil
import json
import math
from sklearn import preprocessing
import sys
import os
import glob
import openpyxl
import warnings
warnings.filterwarnings("ignore") # to have clean looking notebook

material="N87"

basepath="."


nparams = 24



# Model run time for log folders
now = datetime.datetime.now()
nowstr = now.strftime("%Y-%m-%d_%H-%M-%S")
logpath = f"{basepath}/logs/{nowstr}"
isExist = os.path.exists(logpath)
if not isExist:
    os.makedirs(logpath)

os.mkdir('{}/{}'.format(logpath,"script"))
files = glob.glob("." + '/*.py')

# Copy each file to the destination directory
for file in files:
    shutil.copy(file,f"{logpath}/script")

class EarlyStopping:
    def __init__(self, patience=5, min_delta=0):
        self.patience = patience
        self.min_delta = min_delta
        self.counter = 0
        self.best_loss = None
        self.early_stop = False

    def __call__(self, val_loss):
        if self.best_loss is None:
            self.best_loss = val_loss
        elif self.best_loss - val_loss >= self.min_delta:
            self.best_loss = val_loss
            self.counter = 0
        elif self.best_loss - val_loss < self.min_delta:
            self.counter += 1
            # print(f"INFO: Early stopping counter {self.counter} of {self.patience}")
            if self.counter >= self.patience:
                print("INFO: Early stopping")
                self.early_stop = True
def MagNetStats(data,material = "3C94"):
    """
    Plot graphs that are similar to ones required by competition.
    Will need some tweaking and embedding in a pdf
    """
    plt.clf()
    plt.hist(data, weights=np.ones(len(data)) / len(data), bins=50)
    # Highlight the 95th percentile and max
    plt.axvline(np.percentile(data, 95), color='red', linestyle='dashed')
    plt.axvline(data.max(), color='red', linestyle='dashed')
    plt.axvline(data.mean(), color='red', linestyle='dashed')

    x95 = np.percentile(data, 95)
    xmax = data.max()
    labelH = plt.ylim()[1]/2
    xmean = data.mean()
    x99 = np.percentile(data, 99)
    # Annotate the lines with text

    plt.annotate('95 percentile, {:0.2f}%'.format(x95), xy=(x95,plt.ylim()[1]/2), xytext=(np.percentile(data, 95) + 0.01, labelH), color='red')
    plt.annotate('max, {:0.2f}'.format(xmax), xy=(data.max(), 0.001), xytext=(xmax + 0.01,labelH/2), color='green')
    plt.annotate('mean, {:0.2f}'.format(xmean), xy=(xmean, 0.001), xytext=(xmean + 0.01,labelH*1.8), color='green')

    # Show the plot
    plt.title("Error Distribution for {}\n Avg={:0.3f} %, 95-Pct: {:.2f}%, 99-Prc={:0.2f}%, Max = {:0.2f}%".format(material, xmean,x95,x99, xmax))
    # plt
    # plt.show()
    plt.xlabel("Relative Error of Core Loss [%]")
    plt.ylabel("Ratio of Data Points")

    return plt


class Net(nn.Module):
    def __init__(self):
        super(Net, self).__init__()
        # Define a fully connected layers model with three inputs (frequency,T, Bfft(complex numbers)) and one output (power loss).
        self.layers = nn.Sequential(
            nn.Linear(nparams+5,nparams+5),
            nn.ReLU(),
            nn.Linear(nparams+5,15),
            nn.ReLU(),
            nn.Linear(15,15),
            nn.ReLU(),
            nn.Linear(15, 1),
        )

    def forward(self, x):
        """
        """
        return self.layers(x)
    def ReluOut(x):
        pass
        return nn.ReLU()
def count_parameters(model):
    """Count number of parameters for model comparisons
    """
    return sum(p.numel() for p in model.parameters() if p.requires_grad)

def printe(text):
    print(text)
    st.write(text)
def rotate2D(y, degrees,fast=True):
    points = int(degrees/360*y.shape[1])
    array = []

    part2 = y[:, :-points]
    part1 = y[:, -points:]
    if fast:
        return  np.concatenate([part1,part2], axis=1)

    for  i in range(points, len(y)+points):
        array.append(y[i-len(y)])
    return np.array(array)

def tempEncoder(T):
    """
    For consistent Temperature encoding.
        + Ensure one encoder for validation and test training set
    **To change in future away from onehot encoding it.
    """
    T = T.values
    # Encode Temp as 4 variables
    enc = preprocessing.OneHotEncoder()

    # 2. FIT
    enc.fit(T)

    # 3. Transform
    Tlabels = enc.transform(T).toarray()
def get_dataset_pds(B,F,T,VL):
    """Data loader for values.
        material input (folder name): example: "N87"
        basefolder input(main folder a material folder): example: data
    """
    variants=["nov-1","dec-31","ecce-2025"]
    variant="dec-31"
    assert variant in variants, "not in variants"     
    resampled=True
    Toriginal=False
    Freq = F.values
    Flux  = B.values*1024/24
    Power = VL.values
    shifts = False
    T = T.values
    enc = preprocessing.OneHotEncoder()

    if variant=="nov-1": # np.abs(fft) variant
        enc.fit(T)
        Tlabels = enc.transform(T).toarray()
        T = Tlabels
    elif variant=="dec-31": # portable variant
        enc.fit(T)
        Tlabels = enc.transform(T).toarray()
        T = Tlabels
        print(f"Using {variant} variant")
    elif variant=="ecce-2025": # temperature fix variant
        T = np.concatenate((T/100,T**2/100**2,T**3/100**3,T**4/100**4),axis=1)

    if not resampled:
        Flux = scipy.signal.resample(B.values,nparams, axis=1)
        Fluxx = scipy.signal.resample(Flux, 1024, axis=1)
        Error = np.sum(np.abs(B.values-Fluxx),axis=1)/np.sum(np.abs(B.values),axis=1)

    Flux = np.abs(Flux)
    Flux, Power, Freq= np.log10(Flux), np.log10(Power),np.log10(Freq)

    # Reshape data
    Freq = Freq.reshape((-1,1))
    Flux = Flux.reshape((-1,nparams))
    T = T.reshape((-1,4))

    if shifts:
        for phase in [40,80,160]:

            temp1 = np.concatenate((Freq,rotate2D(Flux,phase),T),axis=1)

            temp = np.concatenate((Freq,Flux,T),axis=1)

            temp = np.concatenate((temp,temp1), axis=0)
            Power = np.concatenate((Power,Power), axis=0)
            Freq = np.concatenate((Freq,Freq), axis=0)
            T = np.concatenate((T,T), axis=0)
            Flux= np.concatenate((Flux,Flux), axis=0)
    else:
        temp = np.concatenate((Freq,Flux,T),axis=1)

    # log data

    # printe(np.shape(Freq))
    printe(np.shape(Flux))
    try:
        os.mkdir('{}/{}'.format(logpath,material))
    except:
        pass
    np.savetxt('{}/{}/inputs.csv'.format(logpath, material),temp, delimiter= ", ")
    np.savetxt('{}/{}/outputs.csv'.format(logpath, material), Power, delimiter= ", ")
    if not resampled:
      np.savetxt('{}/{}/RecError.csv'.format(logpath, material), Error, delimiter= ", ")


    in_tensors = torch.from_numpy(temp).view(-1, nparams + 1 +T.shape[1])
    out_tensors = torch.from_numpy(Power).view(-1, 1)


    return torch.utils.data.TensorDataset(in_tensors, out_tensors)


# Config the model training

def main_pd(B,Freq,T,VL,pretrainer="N87",pretrain=True,patience=50,hyper={}):
    """
    Train and test sequence for a single material.
    """
    material=""
    #   Reproducibility
    random.seed(1)
    np.random.seed(1)
    torch.manual_seed(1)

    # Hyperparameters
    NUM_EPOCH = 1000
    BATCH_SIZE = hyper.get("BATCH_SIZE",128)
    DECAY_EPOCH = 100
    DECAY_RATIO = 0.5
    LR_INI = 0.025
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")


    # Load dataset
    dataset = None
    dataset = get_dataset_pds(B,Freq,T,VL)
    # Split the dataset
    train_size = int(0.6* len(dataset))
    # valid_size = len(dataset) - train_size
    valid_size = int(0.2* len(dataset))

    test_size = len(dataset) - train_size - valid_size ## to be loaded seperated
    train_dataset, valid_dataset, test_dataset = torch.utils.data.random_split(dataset, [train_size, valid_size, test_size])
    kwargs = {}
    train_loader = torch.utils.data.DataLoader(train_dataset, batch_size=BATCH_SIZE, shuffle=True, **kwargs)
    valid_loader = torch.utils.data.DataLoader(valid_dataset, batch_size=BATCH_SIZE, shuffle=False, **kwargs)
    test_loader = torch.utils.data.DataLoader(test_dataset, batch_size=BATCH_SIZE, shuffle=False, **kwargs)

    # Setup network
    net = Net().double().to(device) # to avoid reseting parameters?, can be commented out
    pretrainedModel = f"{os.getcwd()}\\pretrained\\Model{pretrainer}.sd"
    print(os.path.isfile(pretrainedModel))
    st.write("pretrainer" , os.path.isfile(pretrainedModel), pretrainedModel )
    st.write(os.getcwd())

    if os.path.isfile(pretrainedModel) and pretrain==True:
        net.load_state_dict(torch.load(pretrainedModel)) # Model parameter initialize
        net.eval()
        i = 0
        for param in net.parameters():
            if i<2:
                param.requires_grad = False
            i += 1
        param.requires_grad = True
        printe(f"Loaded Master net {pretrainer}")
    # Log the number of parameters
    printe(f"Number of parameters: {count_parameters(net)}")
    # Setup optimizer
    criterion = nn.MSELoss()
    optimizer = optim.Adam(net.parameters(), lr=LR_INI)

    # Train the network
    oldEpoch_loss = 10e8
    early_stop=EarlyStopping(patience)
    epoch_valid_loss = 0
    last_epoch=0
    losses={"train":[],
            "val":[]}
    for epoch_i in range(NUM_EPOCH):

        # Train for one epoch
        epoch_train_loss = 0
        net.train()
        optimizer.param_groups[0]['lr'] = LR_INI* (DECAY_RATIO ** (0+ epoch_i // DECAY_EPOCH))

        for inputs, labels in train_loader:
            optimizer.zero_grad()
            outputs = net(inputs.to(device))
            loss = criterion(outputs, labels.to(device))
            loss.backward()
            optimizer.step()

            epoch_train_loss += loss.item()

        epoch_valid_loss = 0
        # Compute Validation Loss
        with torch.no_grad():
            for inputs, labels in valid_loader:
                outputs = net(inputs.to(device))
                loss = criterion(outputs, labels.to(device))

                epoch_valid_loss += loss.item()
        early_stop(epoch_valid_loss)
        losses['train'].append(np.sqrt(epoch_train_loss)/len(train_dataset)*1e5)
        losses['val'].append(np.sqrt(epoch_valid_loss)/ len(valid_dataset)*1e5)
        if early_stop.early_stop:
            break
        if (epoch_i+1)%50 == 0:
          # print(f"Epoch {epoch_i+1:2d} "
          #       f"Train {epoch_train_loss / len(train_dataset) * 1e5:.5f} "
          #       f"Valid {epoch_valid_loss / len(valid_dataset) * 1e5:.5f}")
          newEpoch_loss = epoch_valid_loss / len(valid_dataset)
        
          change = (newEpoch_loss-oldEpoch_loss)/oldEpoch_loss*100
          oldEpoch_loss = newEpoch_loss
          printe(f"Epoch {epoch_i+1:2d} "
                + f"Train {epoch_train_loss / len(train_dataset) * 1e5:.5f} "
                + f"Valid {epoch_valid_loss / len(valid_dataset) * 1e5:.5f} "
                + f"Change {change:.2f}")
          torch.save(net.state_dict(),f"{logpath}/{material}/Model{material}_{epoch_i+1}.sd")
    # model saving

    # onnx 
    # ex_inputs = torch.randn(1, 1,29).to(device)
    # onnx_ = torch.onnx.export(net, ex_inputs)
    # onnx_.save(f"{logpath}/{material}/model.onnx")
    # # entire net

    # Evaluation
    net.eval()
    y_meas = []
    y_pred = []
    with torch.no_grad():
        for inputs, labels in test_loader:
            y_pred.append(net(inputs.to(device)))
            y_meas.append(labels.to(device))

    y_meas = torch.cat(y_meas, dim=0)
    y_pred = torch.cat(y_pred, dim=0)
    printe(f"Test Loss: {F.mse_loss(y_meas, y_pred).item() / len(test_dataset) * 1e5:.5f}")

    yy_pred = 10**(y_pred.cpu().numpy())
    yy_meas = 10**(y_meas.cpu().numpy())
    Error_re = abs(yy_pred-yy_meas)/abs(yy_meas)*100
    Error_re_avg = np.mean(Error_re)
    percentile95_error = np.percentile(Error_re,95)
    Error_re_rms = np.sqrt(np.mean(Error_re ** 2))
    Error_re_max = np.max(Error_re)
    printe(f"Relative Error: {Error_re_avg:.8f}")
    printe(f"RMS Error: {Error_re_rms:.8f}")
    printe(f"MAX Error: {Error_re_max:.8f}")
    printe(f"95 % error Error: {percentile95_error:.8f}")
    fig = plt.figure()
    plotx = MagNetStats(Error_re, material)
    plotx.savefig('{}/{}/Error.png'.format(logpath, material))
    st.pyplot(fig)
    fig = plt.figure()
    plt.semilogy(losses['train'],label="training losses")
    plt.semilogy(losses['val'],label="validation loss")
    plt.xlabel("epoch")
    plt.ylabel("loss")
    plt.title("Loss vs Epoch")
    plt.grid()
    plt.legend()
    plt.savefig('{}/{}/ValidationLoss.png'.format(logpath, material))

    st.pyplot(fig)
    dataFrame = {"meas":yy_meas.flatten(), "pred":yy_pred.flatten(),"error":Error_re.flatten()}

    data = pd.DataFrame(dataFrame)
    ## Reduce data size of log
    data['meas'] = data['meas'].astype('int')
    data['pred'] = data['pred'].astype('int')
    data['error'] = data['error'].round(1)
    data.to_csv("{}/{}_last_stats.csv".format(logpath,material), index=False)


    # save excel format

    # myworkbook=openpyxl.create_sheet(f"{logpath}/{material}/model.xlsx")
    # myworkbook=openpyxl.Workbook()
    myworkbook=openpyxl.load_workbook("base.xlsx")
    worksheet= myworkbook.get_sheet_by_name("CoreLoss")
    sheet = worksheet
    sheet.cell(row=1, column=3).value = "Material XX"
    sheet.cell(row=1, column=5).value = "Author"
    sheet.cell(row=1, column=6).value = "EH"
    sheet.cell(row=1, column=7).value = f"{nowstr}"
    if True:  
        layer_number=0
        for param in net.parameters():
            data=param.data.numpy()
            if len(data.shape)==2:
                columns = [str(i) for i in range(len(data))]
            else:
                columns=["1"]
            if layer_number%2==0:
                sheet_name=f'L{layer_number//2+1}W'
            else:
                sheet_name=f"L{layer_number//2+1}off"
            # worksheet=myworkbook.create_sheet(sheet_name)
            worksheet= myworkbook.get_sheet_by_name(sheet_name)
            sheet = worksheet
            for row in range(1,len(data)+1):
                if len(data.shape) >=2:
                    columns=data.shape[1]
                    for column in range(1,columns+1):
                        sheet.cell(row=row, column=column).value = data[row-1][column-1]
                else:
                    column=1
                    value=data[row-1]
                    sheet.cell(row=row, column=column).value = data[row-1]
            layer_number+=1
    myworkbook.save(f"{logpath}/{material}/model_{nowstr}.xlsx")
    worksheet= myworkbook.get_sheet_by_name("CoreLoss")
    sheet = worksheet
    sheet.cell(row=1, column=3).value = "Material XX"
    sheet.cell(row=1, column=5).value = "Author"
    sheet.cell(row=1, column=6).value = "EH"
    sheet.cell(row=1, column=7).value = f"{nowstr}"
    myworkbook.save(f"{logpath}/{material}/model_{nowstr}.xlsx")

    # save html format


    # save no pytorch format


    # Save the model parameters
    model_save_location=f"{logpath}/Model{material}.sd"
    torch.save(net.state_dict(),model_save_location)
    printe(f"Training finished! Model is saved! {model_save_location}")
    # pickle the model


    torch.save(net, f"{logpath}/{material}/model.pt") # net


    return data, f"{logpath}/Model{material}.sd",f"{logpath}/{material}/model_{nowstr}.xlsx"
def modelRUN():
    global net
    global pretrain
    global logpath
    """
    Loop through all materials.
        1. call main for each material
    """

    old_stdout = sys.stdout

    log_file = open("{}/{}.csv".format(logpath,"logs.log"),"a+")
    if not os.path.exists(logpath):
        os.makedirs(logpath)
    # sys.stdout = log_file
    now = datetime.datetime.now()
    print(now)

    pretrained_materials = ["3E6", "3F4","77","78","N27", "N30","N49", "N87", "3C90", "3C94"]
    pretrained_materials = ["N87","3E6", "3F4","77","78","N27", "N30","N49", "3C90", "3C94"]
    # pretrained_materials = ["Nxx"]
    materials = ["Material A", "Material B", "Material C", "Material D","Material E","N87","3E6", "3F4","77","78","N27", "N30","N49", "3C90", "3C94"]
    Outside = False  #If true I would  be doing pre-training or transfer learning

    # Config the model training
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    data=None
    pretrainer=""

    pretrain = False
    if pretrain:
      for pretrainer in pretrained_materials:
          net = Net().double().to(device)
          torch.set_num_threads(10)
          net = torch.nn.DataParallel(net)
          logpath = f"{basepath}/logs/{nowstr}/{pretrainer}"
          isExist = os.path.exists(logpath)
          if not isExist:
              os.mkdir(logpath)
          for material in materials:
              main(material,pretrainer)
              print(materials)
    else:
        for material in materials:
              main(material,pretrainer)
              print(materials)


if __name__ == "__main__":
    # shutil.copyfile(__file__, '{}/model.py'.format(logpath))
    # modelRUN()
    pass

